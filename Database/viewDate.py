import openpyxl
print('estacion:')
estacion=input()
print('numero de hojas')
n=int(input())
for iloop in range(1,n+1):
    wb=openpyxl.load_workbook(f'{estacion}{iloop}.xlsx')
    ws=wb['datos']
    nf=ws.max_row
    print(ws.cell(2,1).value,' / ',ws.cell(2,2).value,' / ',ws.cell(2,3).value,'  ---  ',ws.cell(nf,1).value,' / ',ws.cell(nf,2).value,' / ',ws.cell(nf,3).value)
    wb.close()
'''wb=openpyxl.load_workbook(f'{estacion}.xlsx')
ws=wb.active
nf=ws.max_row
print('Hoja resultante')
print(ws.cell(2,1).value,' / ',ws.cell(2,2).value,' / ',ws.cell(2,3).value,'  ---  ',ws.cell(nf,1).value,' / ',ws.cell(nf,2).value,' / ',ws.cell(nf,3).value)
wb.close()'''

